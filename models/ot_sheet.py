# -*- coding: utf-8 -*-
import base64
import io
import calendar
from datetime import date
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class HrOtSheet(models.Model):
    _name = 'hr.ot.sheet'
    _description = 'OT Sheet Header'

    # ------------------------------------------------------------------
    # header fields
    # ------------------------------------------------------------------
    name = fields.Char(string='Reference', copy=False, index=True)
    month = fields.Selection([
        ('1', 'January'), ('2', 'February'), ('3', 'March'),
        ('4', 'April'), ('5', 'May'), ('6', 'June'),
        ('7', 'July'), ('8', 'August'), ('9', 'September'),
        ('10', 'October'), ('11', 'November'), ('12', 'December')
    ], required=True)
    year = fields.Integer(required=True, default=lambda _: date.today().year)
    line_ids = fields.One2many('hr.ot.sheet.line', 'sheet_id', string='OT Lines')
    state = fields.Selection([('draft', 'Draft'), ('done', 'Done')], default='draft')

    import_file = fields.Binary(string='OT Excel file')
    import_filename = fields.Char()

    # ------------------------------------------------------------------
    # sequence for name 001/10/2025
    # ------------------------------------------------------------------
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', 'New') == 'New':
                month = int(vals.get('month'))
                year = int(vals.get('year'))
                seq = self.env['ir.sequence'].next_by_code('hr.ot.sheet') or '001'
                vals['name'] = f"{seq}/{month:02d}/{year}"
        return super().create(vals_list)

    # ------------------------------------------------------------------
    # import Excel button (reads hours)
    # ------------------------------------------------------------------
    def action_import_excel(self):
        self.ensure_one()
        if not self.import_file:
            raise UserError(_('Please choose an Excel file first.'))
        if not load_workbook:
            raise UserError(_('python library "openpyxl" is missing on the server.'))

        data = base64.b64decode(self.import_file)
        ws = load_workbook(io.BytesIO(data), data_only=True).active

        error_lines, created = [], 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            row = list(row) + [None] * 6
            emp_code, emp_name, ot_normal_hrs, ot_holiday_hrs, late_ded_hrs, description = row[:6]

            try:
                ot_normal_hrs = float(ot_normal_hrs or 0)
                ot_holiday_hrs = float(ot_holiday_hrs or 0)
                late_ded_hrs = float(late_ded_hrs or 0)
            except Exception:
                error_lines.append((idx, 'Invalid numeric value'))
                continue

            employee = None
            if emp_code:
                employee = self.env['hr.employee'].search([('barcode', '=', str(emp_code))], limit=1)
            if not employee and emp_name:
                employee = self.env['hr.employee'].search([('name', 'ilike', str(emp_name))], limit=1)
            if not employee:
                error_lines.append((idx, f'Employee not found: {emp_code}/{emp_name}'))
                continue

            self.env['hr.ot.sheet.line'].create({
                'sheet_id': self.id,
                'employee_id': employee.id,
                'ot_normal_hrs': ot_normal_hrs,
                'ot_holiday_hrs': ot_holiday_hrs,
                'late_ded_hrs': late_ded_hrs,
                'description': description or '',  # allow manual override
            })
            created += 1

        self.import_file = False
        message = f'Imported {created} rows.'
        if error_lines:
            message += ' Errors: ' + ', '.join([f'Row {r}: {m}' for r, m in error_lines])
        # reload form
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'main',
            'flags': {'reload': True},
        }

    # ------------------------------------------------------------------
    # apply to payslips button
    # ------------------------------------------------------------------
    def action_apply_to_payslips(self):
        self.ensure_one()
        Payroll = self.env['hr.payslip']
        Input = self.env['hr.payslip.input']
        InputType = self.env['hr.payslip.input.type']

        codes = ('OT_NORMAL', 'OT_HOLIDAY', 'LATE_DEDUCTION')
        input_types = {
            c: InputType.search([('code', '=', c)], limit=1) or
               InputType.create({'name': c.replace('_', ' ').title(), 'code': c})
            for c in codes
        }

        for line in self.line_ids:
            year = int(self.year)
            month = int(self.month)
            last_day = calendar.monthrange(year, month)[1]
            date_from = date(year, month, 1)
            date_to = date(year, month, last_day)

            slip = Payroll.search([
                ('employee_id', '=', line.employee_id.id),
                ('date_from', '<=', date_to),
                ('date_to', '>=', date_from),
            ], limit=1)

            if not slip:
                contract = self.env['hr.contract'].search([
                    ('employee_id', '=', line.employee_id.id),
                    ('state', 'in', ('open', 'close'))
                ], limit=1)
                if not contract:
                    continue
                struct = (contract.structure_type_id.default_struct_id or
                          self.env['hr.payroll.structure'].search([], limit=1))
                if not struct:
                    raise UserError(_('No salary structure for employee %s') % line.employee_id.name)
                slip = Payroll.create({
                    'employee_id': contract.employee_id.id,
                    'contract_id': contract.id,
                    'struct_id': struct.id,
                    'date_from': date_from,
                    'date_to': date_to,
                    'name': self.env['hr.ot.import.wizard']._generate_payslip_name(
                        contract.employee_id, struct.id, date_from, date_to),
                })

            def upsert(code, amount, desc=None):
                if not amount:
                    return
                existing = Input.search([
                    ('payslip_id', '=', slip.id),
                    ('input_type_id', '=', input_types[code].id)
                ], limit=1)
                if existing:
                    existing.write({'amount': amount, 'name': desc or ''})
                else:
                    Input.create({
                        'payslip_id': slip.id,
                        'input_type_id': input_types[code].id,
                        'amount': amount,
                        'name': desc or '',
                    })

            upsert('OT_NORMAL', line.ot_normal, line.description)
            upsert('OT_HOLIDAY', line.ot_holiday, line.description)
            upsert('LATE_DEDUCTION', -abs(line.late_deduction) if line.late_deduction else 0, line.description)

            line.applied = True

        self.state = 'done'
        return {'type': 'ir.actions.act_window_close'}


class HrOtSheetLine(models.Model):
    _name = 'hr.ot.sheet.line'
    _description = 'OT Sheet Line'

    sheet_id = fields.Many2one('hr.ot.sheet', ondelete='cascade')
    employee_id = fields.Many2one('hr.employee', string='Employee', required=True)

    # hours entered by user
    ot_normal_hrs   = fields.Float(string='Normal OT Hours', digits=(4, 2))
    ot_holiday_hrs  = fields.Float(string='Holiday OT Hours', digits=(4, 2))
    late_ded_hrs    = fields.Float(string='Late Deduction Hours', digits=(4, 2))

    # rates from employee master
    ot_rate         = fields.Float(related='employee_id.x_studio_ot_hours', string='OT Rate', readonly=True)
    employee_rate   = fields.Float(related='employee_id.x_studio_employee_rate', string='Employee Rate', readonly=True)

    # computed amounts
    ot_normal       = fields.Monetary(string='Normal OT Amount', currency_field='company_currency',
                                      compute='_compute_amounts', store=True)
    ot_holiday      = fields.Monetary(string='Holiday OT Amount', currency_field='company_currency',
                                      compute='_compute_amounts', store=True)
    late_deduction  = fields.Monetary(string='Late Deduction Amount', currency_field='company_currency',
                                      compute='_compute_amounts', store=True)

    company_currency = fields.Many2one('res.currency', related='employee_id.company_id.currency_id', readonly=True)
    applied = fields.Boolean(default=False)

    description = fields.Char(string='Description', compute='_compute_description', store=True)

    @api.depends('ot_normal_hrs', 'ot_holiday_hrs', 'late_ded_hrs', 'ot_rate', 'employee_rate')
    def _compute_amounts(self):
        for rec in self:
            rec.ot_normal      = rec.ot_normal_hrs * rec.ot_rate
            rec.ot_holiday     = rec.ot_holiday_hrs * rec.ot_rate
            rec.late_deduction = rec.late_ded_hrs * rec.employee_rate

    @api.depends('ot_normal_hrs', 'ot_holiday_hrs', 'late_ded_hrs', 'ot_rate', 'employee_rate')
    def _compute_description(self):
        for rec in self:
            desc = []
            if rec.ot_normal_hrs:
                desc.append(f"{rec.ot_normal_hrs} OT hours @ {rec.ot_rate}")
            if rec.ot_holiday_hrs:
                desc.append(f"{rec.ot_holiday_hrs} OT hours @ {rec.ot_rate}")
            if rec.late_ded_hrs:
                desc.append(f"{rec.late_ded_hrs} Late hours @ {rec.employee_rate}")
            rec.description = ' | '.join(desc) if desc else ''
import base64
import io
import calendar
from datetime import date
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class ImportOTWizard(models.TransientModel):
    _name = 'hr.ot.import.wizard'
    _description = 'Import OT from Excel'

    file = fields.Binary('OT Excel File', required=True)
    filename = fields.Char('File Name')
    month = fields.Selection([(str(i), str(i)) for i in range(1, 13)], string='Month', required=True)
    year = fields.Integer('Year', required=True)
    sheet_ref = fields.Many2one('hr.ot.sheet', string='OT Sheet')

    def _check_openpyxl(self):
        if not load_workbook:
            raise UserError(_('openpyxl is not installed on the server. Please install it (pip install openpyxl)'))

    def action_import(self):
        self._check_openpyxl()
        if not self.file:
            raise UserError(_('Please upload a file.'))

        data = base64.b64decode(self.file)
        fp = io.BytesIO(data)
        wb = load_workbook(fp, data_only=True)
        ws = wb.active

        # Expected columns: Employee Code | Employee Name | Normal OT | Holiday OT | Late Deduction
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        sheet = self.sheet_ref or self.env['hr.ot.sheet'].create({
            'name': f'OT/{self.year}/{self.month}',
            'month': self.month,
            'year': self.year
        })

        error_lines = []
        created = 0
        for idx, row in enumerate(rows, start=2):
            row = list(row) + [None]*5
            emp_code, emp_name, ot_normal, ot_holiday, late_deduction = row[:5]

            try:
                ot_normal = float(ot_normal or 0)
                ot_holiday = float(ot_holiday or 0)
                late_deduction = float(late_deduction or 0)
            except Exception:
                error_lines.append((idx, 'Invalid numeric value'))
                continue

            employee = None
            if emp_code:
                employee = self.env['hr.employee'].search([('barcode', '=', str(emp_code))], limit=1)
            if not employee and emp_name:
                employee = self.env['hr.employee'].search([('name', 'ilike', str(emp_name))], limit=1)
            if not employee:
                error_lines.append((idx, f'Employee not found: {emp_code}/{emp_name}'))
                continue

            self.env['hr.ot.sheet.line'].create({
                'sheet_id': sheet.id,
                'employee_id': employee.id,
                'ot_normal': ot_normal,
                'ot_holiday': ot_holiday,
                'late_deduction': late_deduction,
            })
            created += 1

        msg = f'Imported {created} rows.'
        if error_lines:
            msg += ' Errors: ' + ', '.join([f'Row {r}: {m}' for r, m in error_lines])

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'hr.ot.sheet',
            'view_mode': 'form',
            'res_id': sheet.id,
        }

    def action_create_inputs(self):
        if not self.sheet_ref:
            raise UserError(_('Select an OT sheet to apply.'))
        sheet = self.sheet_ref
        Payroll = self.env['hr.payslip']
        Input = self.env['hr.payslip.input']

        for line in sheet.line_ids:
            year = int(sheet.year)
            month = int(sheet.month)
            last_day = calendar.monthrange(year, month)[1]
            date_from = date(year, month, 1)
            date_to = date(year, month, last_day)

            payslip = Payroll.search([
                ('employee_id', '=', line.employee_id.id),
                ('date_from', '>=', date_from),
                ('date_to', '<=', date_to),
            ], limit=1)

            if not payslip:
                contract = self.env['hr.contract'].search([
                    ('employee_id', '=', line.employee_id.id),
                    ('state', 'in', ('open', 'close'))
                ], limit=1)

                if not contract:
                    continue

                # Odoo 17: get salary structure from contract.structure_type_id.default_struct_id
                struct_id = False
                if contract.structure_type_id and contract.structure_type_id.default_struct_id:
                    struct_id = contract.structure_type_id.default_struct_id.id
                else:
                    # fallback: pick any salary structure in company
                    struct = self.env['hr.payroll.structure'].search([], limit=1)
                    struct_id = struct.id if struct else False

                if not struct_id:
                    raise UserError(_('No salary structure found for employee %s (%s). Please assign a salary structure first.') %
                                    (contract.employee_id.name, contract.employee_id.barcode))

                payslip = Payroll.create({
                    'employee_id': contract.employee_id.id,
                    'contract_id': contract.id,
                    'struct_id': struct_id,
                    'date_from': date_from,
                    'date_to': date_to,
                    'state': 'draft',
                })

            def upsert_input(payslip, code, amount):
                input_type = self.env['hr.payslip.input.type'].search([('code', '=', code)], limit=1)
                if not input_type:
                    input_type = self.env['hr.payslip.input.type'].create({'name': code, 'code': code})

                existing = Input.search([('payslip_id', '=', payslip.id), ('input_type_id', '=', input_type.id)], limit=1)
                if existing:
                    existing.write({'amount': amount})
                else:
                    Input.create({
                        'payslip_id': payslip.id,
                        'input_type_id': input_type.id,
                        'amount': amount,
                    })

            upsert_input(payslip, 'OT_NORMAL', line.ot_normal)
            upsert_input(payslip, 'OT_HOLIDAY', line.ot_holiday)
            if line.late_deduction:
                upsert_input(payslip, 'LATE_DEDUCTION', -abs(line.late_deduction))

            line.applied = True

        sheet.state = 'done'
        return {'type': 'ir.actions.act_window_close'}